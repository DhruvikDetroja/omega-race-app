import streamlit as st
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image

# === Coefficient Table for 2nd Week of April ===
coefficient_data = [
    (0, 4999, 0),
    (5000, 5099, 0.85),
    (5100, 5199, 0.875),
    (5200, 5299, 0.9),
    (5300, 5399, 0.925),
    (5400, 5499, 0.95),
    (5500, 5599, 0.975),
    (5600, 5699, 1),
    (5700, 5799, 1),
    (5800, 5899, 1),
    (5900, 5999, 1),
    (6000, 6999, 1),
]

def parse_player_data(file_contents):
    lines = file_contents.strip().split('\n')[1:]  # Skip header
    players = []
    for line in lines:
        match = re.match(r".*?(\d{4})\s+(.*)", line.strip())
        if match:
            trophy = int(match.group(1))
            name = match.group(2).strip()
            players.append((name, trophy))
    return players

def generate_excel(players):
    wb = Workbook()
    ws = wb.active
    ws.title = "Omega Race"

    headers = [
        "Name", "Trophies",
        "1st Attack Stars", "1st Attack destruction %",
        "2nd Attack Stars", "2nd Attack destruction %",
        "3rd Attack Stars", "3rd Attack destruction %",
        "Total Destruction", "Coefficient", "Coefficient Destruction", "Final Score"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for idx, (name, trophy) in enumerate(players, start=2):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=trophy)
        ws.cell(row=idx, column=9, value=f"=SUM(D{idx},F{idx},H{idx})")
        ws.cell(row=idx, column=10, value=f"=IFERROR(VLOOKUP(B{idx},Sheet2!A:C,3,TRUE),1)")
        ws.cell(row=idx, column=11, value=f"=I{idx}*J{idx}")
        ws.cell(row=idx, column=12, value=f"=K{idx} + (IF(C{idx}=3,5,0) + IF(E{idx}=3,5,0) + IF(G{idx}=3,5,0))")

    # Coefficient table in Sheet2
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Trophy Min", "Trophy Max", "Coefficient"])
    for row in coefficient_data:
        ws2.append(row)

    output = BytesIO()
    wb.save(output)
    return output

# ===================== STREAMLIT UI ========================

st.markdown(
    """
    <div style='text-align: center;'>
        <img src='https://raw.githubusercontent.com/DhruvikDetroja/omega-race-app/main/logo.png' width='200'>
    </div>
    """, unsafe_allow_html=True
)


st.set_page_config(page_title="Omega Race Sheet Generator", layout="centered")
st.title("🏆 Omega Race Sheet Generator (Week 2 Only)")
st.write("Upload your `Player_Data.txt` and download the formatted Excel sheet.")
st.write("Built by Dhruvik aka `Elephant Fart Lol`")

uploaded_file = st.file_uploader("Upload Player_Data.txt", type="txt")

if uploaded_file:
    file_contents = uploaded_file.read().decode("utf-8")
    players = parse_player_data(file_contents)
    if not players:
        st.warning("No valid player data found.")
    else:
        excel_output = generate_excel(players)
        st.success("✅ Excel file generated successfully!")
        st.download_button(
            label="📥 Download Excel File",
            data=excel_output.getvalue(),
            file_name="Omega_Race_sheet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
